"""
Advanced Excel extraction services for both .xlsx and .xls formats
Supports: merged cells, images, charts, formatting, formulas, and more
"""
import io
import base64
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, Any, Optional, List
from datetime import datetime
from .structured_logger import logger, OperationTimer
from .file_validator import validate_file


class ExcelExtractor:
    """
    Extracts text and advanced features from Excel files with full fidelity.
    Supports: .xlsx and .xls formats, merged cells, images, charts, formatting.
    """
    
    def __init__(self, extract_images: bool = True, extract_charts: bool = True, 
                 extract_formatting: bool = True):
        """
        Initialize extractor with feature flags.
        
        Args:
            extract_images: Whether to extract embedded images
            extract_charts: Whether to extract chart information
            extract_formatting: Whether to extract cell formatting
        """
        self.extract_images = extract_images
        self.extract_charts = extract_charts
        self.extract_formatting = extract_formatting
    
    def extract_from_bytes(self, file_bytes: bytes, filename: str, request_id: str = None) -> Dict[str, Any]:
        """
        Extract Excel content with full fidelity from bytes.
        
        Args:
            file_bytes: Excel file content as bytes
            filename: Name of the file (for format detection)
            request_id: Optional request ID for tracing
            
        Returns:
            Dictionary containing extracted data with structure:
            {
                "filename": str,
                "sheets": [
                    {
                        "name": str,
                        "data": {
                            "A1": [value, type, formatting],  # formatting is optional
                        },
                        "merged_cells": [...],  # if any
                        "images": [...],  # if extract_images=True
                        "charts": [...],  # if extract_charts=True
                        "column_widths": {...},
                        "row_heights": {...},
                        "summary": {...}
                    }
                ],
                "metadata": {...}
            }
        """
        with OperationTimer("excel_extraction", filename=filename, request_id=request_id):
            try:
                # Validate file
                validation_result = validate_file(
                    filename or "",
                    len(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                if not validation_result["valid"]:
                    logger.log_validation_error(
                        filename=filename or "unknown",
                        errors=validation_result["errors"],
                        request_id=request_id
                    )
                    raise ValueError(f"File validation failed: {validation_result['errors']}")
                
                logger.info(
                    "Starting Excel extraction",
                    filename=filename,
                    file_size=len(file_bytes),
                    extract_images=self.extract_images,
                    extract_charts=self.extract_charts,
                    extract_formatting=self.extract_formatting,
                    request_id=request_id
                )
                
                if filename.lower().endswith(('.xlsx', '.xlsm')):
                    result = self._extract_xlsx(file_bytes, filename, request_id)
                elif filename.lower().endswith('.xls'):
                    result = self._extract_xls(file_bytes, filename, request_id)
                else:
                    raise ValueError(f"Unsupported Excel format: {filename}")
                
                logger.log_file_processing(
                    filename=filename,
                    file_size=len(file_bytes),
                    file_type="excel",
                    operation="extraction",
                    success=True,
                    request_id=request_id
                )
                
                return result
                
            except Exception as e:
                logger.error(
                    "Excel extraction failed",
                    filename=filename,
                    error=str(e),
                    extract_images=self.extract_images,
                    extract_charts=self.extract_charts,
                    extract_formatting=self.extract_formatting,
                    request_id=request_id
                )
                raise ValueError(f"Error extracting Excel data: {str(e)}")
    
    def _extract_xlsx(self, file_bytes: bytes, filename: str, request_id: str = None) -> Dict[str, Any]:
        """Extract from modern Excel format (.xlsx) using openpyxl."""
        # Load workbook with full features
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes), 
            data_only=False,  # Keep formulas
            keep_vba=False
        )
        
        sheets_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = {}
            non_empty_count = 0
            
            # Get actual used range
            max_row = sheet.max_row or 1
            max_col = sheet.max_column or 1
            
            # Extract cell data with formatting
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = self._get_cell_value(cell)
                    data_type = self._get_cell_type(cell)
                    
                    # Convert value to string properly
                    if value is None:
                        value_str = ""
                    elif isinstance(value, datetime):
                        value_str = value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        value_str = str(value)
                    
                    # Skip empty cells
                    if not value_str.strip():
                        continue
                    
                    non_empty_count += 1
                    
                    # Build cell data
                    cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                    cell_data = [value_str, data_type]
                    
                    # Add formatting if enabled
                    if self.extract_formatting:
                        formatting = self._extract_cell_formatting(cell)
                        if formatting:  # Only add if there's actual formatting
                            cell_data.append(formatting)
                    
                    sheet_data[cell_ref] = cell_data
            
            # Build sheet info
            sheet_info = {
                "name": sheet_name,
                "data": sheet_data,
                "summary": {
                    "total_rows": max_row,
                    "total_cols": max_col,
                    "non_empty_cells": non_empty_count
                }
            }
            
            # Extract merged cells
            merged_cells = self._extract_merged_cells(sheet)
            if merged_cells:
                sheet_info["merged_cells"] = merged_cells
            
            # Extract column widths and row heights
            column_widths, row_heights = self._extract_dimensions(sheet, max_col, max_row)
            if column_widths:
                sheet_info["column_widths"] = column_widths
            if row_heights:
                sheet_info["row_heights"] = row_heights
            
            # Extract images
            if self.extract_images:
                images = self._extract_images(sheet)
                if images:
                    sheet_info["images"] = images
            
            # Extract charts
            if self.extract_charts:
                charts = self._extract_charts(sheet)
                if charts:
                    sheet_info["charts"] = charts
            
            # Extract conditional formatting
            if self.extract_formatting and sheet.conditional_formatting:
                cond_fmt = self._extract_conditional_formatting(sheet)
                if cond_fmt:
                    sheet_info["conditional_formatting"] = cond_fmt
            
            # Extract data validation
            if sheet.data_validations:
                data_val = self._extract_data_validations(sheet)
                if data_val:
                    sheet_info["data_validations"] = data_val
            
            sheets_data.append(sheet_info)
        
        # Extract workbook metadata
        metadata = self._extract_workbook_metadata(workbook)
        
        return {
            "filename": filename,
            "sheets": sheets_data,
            "metadata": metadata
        }
    
    def _extract_xls(self, file_bytes: bytes, filename: str, request_id: str = None) -> Dict[str, Any]:
        """Extract from legacy Excel format (.xls) using pandas."""
        excel_file = io.BytesIO(file_bytes)
        
        try:
            # Read all sheets
            xls = pd.ExcelFile(excel_file, engine='xlrd')
            sheets_data = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                sheet_data = {}
                non_empty_count = 0
                
                for row_idx in range(len(df)):
                    for col_idx in range(len(df.columns)):
                        value = df.iloc[row_idx, col_idx]
                        
                        # Convert value to string properly
                        if pd.isna(value):
                            value_str = ""
                        elif isinstance(value, datetime):
                            value_str = value.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            value_str = str(value)
                        
                        # Skip empty cells
                        if not value_str.strip():
                            continue
                        
                        non_empty_count += 1
                        
                        # Store as key-value: cell_ref -> [value, type]
                        cell_ref = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                        sheet_data[cell_ref] = [value_str, self._infer_data_type(value)]
                
                sheets_data.append({
                    "name": sheet_name,
                    "data": sheet_data,
                    "summary": {
                        "total_rows": len(df),
                        "total_cols": len(df.columns),
                        "non_empty_cells": non_empty_count
                    }
                })
            
            return {
                "filename": filename,
                "sheets": sheets_data,
                "metadata": {
                    "format": "xls",
                    "note": "Limited feature extraction for legacy format"
                }
            }
            
        except Exception as e:
            raise ValueError(f"Error reading .xls file: {str(e)}")
    
    def _get_cell_value(self, cell) -> Any:
        """Get the actual value from an openpyxl cell."""
        try:
            # Handle merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                return ""
            
            return cell.value
        except Exception:
            return ""
    
    def _get_cell_type(self, cell) -> str:
        """Determine the data type of a cell."""
        try:
            # Handle merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                return "merged"
            
            if cell.value is None:
                return "empty"
            
            if cell.data_type == 'f':
                return "formula"
            elif cell.is_date:
                return "date"
            elif cell.data_type == 'n':
                return "number"
            elif cell.data_type == 's':
                return "string"
            elif cell.data_type == 'b':
                return "boolean"
            else:
                return "string"
        except Exception:
            return "unknown"
    
    def _extract_cell_formatting(self, cell) -> Optional[Dict[str, Any]]:
        """Extract formatting information from a cell."""
        try:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                return None
            
            formatting = {}
            
            # Font formatting
            if cell.font:
                font_info = {}
                if cell.font.bold:
                    font_info["bold"] = True
                if cell.font.italic:
                    font_info["italic"] = True
                if cell.font.underline and cell.font.underline != 'none':
                    font_info["underline"] = cell.font.underline
                if cell.font.size:
                    font_info["size"] = cell.font.size
                if cell.font.color and cell.font.color.rgb:
                    font_info["color"] = str(cell.font.color.rgb)
                if cell.font.name:
                    font_info["name"] = cell.font.name
                
                if font_info:
                    formatting["font"] = font_info
            
            # Fill/background color
            if cell.fill and cell.fill.start_color:
                if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                    rgb = str(cell.fill.start_color.rgb)
                    # Ignore default white background
                    if rgb != '00000000' and rgb != 'FFFFFFFF':
                        formatting["fill"] = {
                            "fgColor": rgb,
                            "patternType": cell.fill.patternType
                        }
            
            # Alignment
            if cell.alignment:
                align_info = {}
                if cell.alignment.horizontal:
                    align_info["horizontal"] = cell.alignment.horizontal
                if cell.alignment.vertical:
                    align_info["vertical"] = cell.alignment.vertical
                if cell.alignment.wrap_text:
                    align_info["wrap_text"] = True
                if cell.alignment.text_rotation:
                    align_info["text_rotation"] = cell.alignment.text_rotation
                
                if align_info:
                    formatting["alignment"] = align_info
            
            # Border
            if cell.border and any([cell.border.left.style, cell.border.right.style,
                                   cell.border.top.style, cell.border.bottom.style]):
                border_info = {}
                for side in ['left', 'right', 'top', 'bottom']:
                    border_side = getattr(cell.border, side)
                    if border_side and border_side.style:
                        border_info[side] = {
                            "style": border_side.style,
                            "color": str(border_side.color.rgb) if border_side.color and border_side.color.rgb else None
                        }
                
                if border_info:
                    formatting["border"] = border_info
            
            # Number format
            if cell.number_format and cell.number_format != 'General':
                formatting["number_format"] = cell.number_format
            
            # Only return formatting if there's actual formatting data
            return formatting if formatting else None
            
        except Exception as e:
            return None
    
    def _extract_merged_cells(self, sheet) -> List[Dict[str, Any]]:
        """Extract merged cell ranges."""
        merged_cells = []
        
        try:
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "range": str(merged_range),
                    "top_left": {
                        "col": merged_range.min_col,
                        "row": merged_range.min_row
                    },
                    "bottom_right": {
                        "col": merged_range.max_col,
                        "row": merged_range.max_row
                    }
                })
        except Exception:
            pass
        
        return merged_cells
    
    def _extract_dimensions(self, sheet, max_col: int, max_row: int) -> tuple:
        """Extract column widths and row heights."""
        column_widths = {}
        row_heights = {}
        
        try:
            # Column widths
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in sheet.column_dimensions:
                    width = sheet.column_dimensions[col_letter].width
                    if width and width != 8.43:  # Default width
                        column_widths[col_letter] = round(width, 2)
            
            # Row heights
            for row_idx in range(1, max_row + 1):
                if row_idx in sheet.row_dimensions:
                    height = sheet.row_dimensions[row_idx].height
                    if height and height != 15:  # Default height
                        row_heights[row_idx] = round(height, 2)
        except Exception:
            pass
        
        return column_widths, row_heights
    
    def _extract_images(self, sheet) -> List[Dict[str, Any]]:
        """Extract embedded images from sheet."""
        images = []
        
        try:
            for image in sheet._images:
                # Extract anchor as cell reference
                anchor_str = None
                if hasattr(image, 'anchor') and image.anchor:
                    try:
                        # TwoCell Anchor has _from attribute with col and row
                        if hasattr(image.anchor, '_from'):
                            from_cell = image.anchor._from
                            col = get_column_letter(from_cell.col + 1)  # 0-indexed
                            row = from_cell.row + 1  # 0-indexed
                            anchor_str = f"{col}{row}"
                        else:
                            anchor_str = "A1"  # Default position
                    except:
                        anchor_str = "A1"  # Default if parsing fails
                
                img_data = {
                    "anchor": anchor_str,
                    "width": image.width,
                    "height": image.height,
                }
                
                # Convert image to base64
                try:
                    img_bytes = io.BytesIO()
                    if hasattr(image, '_data'):
                        # Try to get image data
                        pil_image = image._data()
                        pil_image.save(img_bytes, format='PNG')
                        img_data["data"] = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
                        img_data["format"] = "PNG"
                except Exception as e:
                    img_data["error"] = f"Could not extract image data: {str(e)}"
                
                images.append(img_data)
        except Exception:
            pass
        
        return images
    
    def _extract_charts(self, sheet) -> List[Dict[str, Any]]:
        """Extract chart information from sheet."""
        charts = []
        
        try:
            for chart in sheet._charts:
                # Extract anchor as cell reference (e.g., "G2")
                anchor_str = None
                if hasattr(chart, 'anchor') and chart.anchor:
                    try:
                        # TwoCell Anchor has _from attribute with col and row
                        if hasattr(chart.anchor, '_from'):
                            from_cell = chart.anchor._from
                            col = get_column_letter(from_cell.col + 1)  # 0-indexed
                            row = from_cell.row + 1  # 0-indexed
                            anchor_str = f"{col}{row}"
                        else:
                            anchor_str = "G2"  # Default position
                    except:
                        anchor_str = "G2"  # Default if parsing fails
                
                chart_info = {
                    "type": chart.__class__.__name__,
                    "anchor": anchor_str,
                }
                
                # Extract title
                title_text = None
                if hasattr(chart, 'title') and chart.title:
                    try:
                        # Try to get title text
                        if hasattr(chart.title, 'tx') and chart.title.tx:
                            if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                                # Extract from rich text
                                if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                                    for para in chart.title.tx.rich.p:
                                        if hasattr(para, 'r') and para.r:
                                            for run in para.r:
                                                if hasattr(run, 't'):
                                                    title_text = run.t
                                                    break
                                        if title_text:
                                            break
                        # Fallback to simple string conversion
                        if not title_text:
                            title_text = str(chart.title)[:100]  # Truncate if too long
                    except:
                        title_text = "Chart"
                
                if title_text:
                    chart_info["title"] = title_text
                
                # Extract series data
                try:
                    series_info = []
                    if hasattr(chart, 'series'):
                        for s in chart.series:
                            series_data = {}
                            
                            # Extract values reference
                            if hasattr(s, 'val') and s.val:
                                try:
                                    # Get the formula string from the Reference
                                    if hasattr(s.val, 'f'):
                                        series_data["values"] = s.val.f
                                    elif hasattr(s.val, 'formula'):
                                        series_data["values"] = s.val.formula
                                except:
                                    pass
                            
                            # Extract categories reference
                            if hasattr(s, 'cat') and s.cat:
                                try:
                                    if hasattr(s.cat, 'f'):
                                        series_data["categories"] = s.cat.f
                                    elif hasattr(s.cat, 'formula'):
                                        series_data["categories"] = s.cat.formula
                                except:
                                    pass
                            
                            # Extract title
                            if hasattr(s, 'title') and s.title:
                                try:
                                    # Try to extract title value
                                    if hasattr(s.title, 'v'):
                                        series_data["title"] = s.title.v
                                except:
                                    pass
                            
                            if series_data:
                                series_info.append(series_data)
                    
                    if series_info:
                        chart_info["series"] = series_info
                except Exception:
                    pass
                
                # Extract axis information
                try:
                    if hasattr(chart, 'x_axis') and chart.x_axis:
                        if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                            chart_info["x_axis_title"] = str(chart.x_axis.title)
                    
                    if hasattr(chart, 'y_axis') and chart.y_axis:
                        if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                            chart_info["y_axis_title"] = str(chart.y_axis.title)
                except Exception:
                    pass
                
                charts.append(chart_info)
        except Exception:
            pass
        
        return charts
    
    def _extract_conditional_formatting(self, sheet) -> List[Dict[str, Any]]:
        """Extract conditional formatting rules."""
        cond_formats = []
        
        try:
            for range_string, rules in sheet.conditional_formatting._cf_rules.items():
                for rule in rules:
                    cond_format = {
                        "range": range_string,
                        "type": rule.type if hasattr(rule, 'type') else None,
                        "priority": rule.priority if hasattr(rule, 'priority') else None
                    }
                    
                    # Add rule-specific details
                    if hasattr(rule, 'formula') and rule.formula:
                        cond_format["formula"] = rule.formula
                    
                    cond_formats.append(cond_format)
        except Exception:
            pass
        
        return cond_formats
    
    def _extract_data_validations(self, sheet) -> List[Dict[str, Any]]:
        """Extract data validation rules."""
        validations = []
        
        try:
            for dv in sheet.data_validations.dataValidation:
                validation = {
                    "ranges": [str(r) for r in dv.cells.ranges] if dv.cells else [],
                    "type": dv.type if hasattr(dv, 'type') else None,
                }
                
                if hasattr(dv, 'formula1') and dv.formula1:
                    validation["formula1"] = dv.formula1
                if hasattr(dv, 'formula2') and dv.formula2:
                    validation["formula2"] = dv.formula2
                if hasattr(dv, 'allow_blank'):
                    validation["allow_blank"] = dv.allow_blank
                if hasattr(dv, 'showDropDown'):
                    validation["show_dropdown"] = dv.showDropDown
                
                validations.append(validation)
        except Exception:
            pass
        
        return validations
    
    def _extract_workbook_metadata(self, workbook) -> Dict[str, Any]:
        """Extract workbook-level metadata."""
        metadata = {
            "format": "xlsx",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames
        }
        
        try:
            if workbook.properties:
                props = workbook.properties
                if props.creator:
                    metadata["creator"] = props.creator
                if props.created:
                    metadata["created"] = props.created.isoformat()
                if props.modified:
                    metadata["modified"] = props.modified.isoformat()
                if props.title:
                    metadata["title"] = props.title
                if props.subject:
                    metadata["subject"] = props.subject
        except Exception:
            pass
        
        return metadata
    
    def _infer_data_type(self, value) -> str:
        """Infer data type for pandas-based extraction."""
        if pd.isna(value) or value == "":
            return "empty"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, datetime):
            return "date"
        elif isinstance(value, str):
            # Try to detect if it's a date
            try:
                pd.to_datetime(value)
                return "date"
            except:
                return "string"
        else:
            return "unknown"
    
    def extract_as_text(self, file_bytes: bytes, filename: str) -> str:
        """
        Extract Excel content and return it in formatted text.
        
        Args:
            file_bytes: Excel file content as bytes
            filename: Name of the file
            
        Returns:
            Formatted text string with all extracted data
        """
        excel_data = self.extract_from_bytes(file_bytes, filename)
        
        # Convert to readable text format
        all_text = []
        all_text.append(f"=" * 80)
        all_text.append(f"EXCEL FILE: {excel_data['filename']}")
        all_text.append(f"=" * 80)
        all_text.append("")
        
        # Metadata
        if "metadata" in excel_data:
            all_text.append("METADATA:")
            for key, value in excel_data["metadata"].items():
                all_text.append(f"  {key}: {value}")
            all_text.append("")
        
        # Sheets
        for sheet in excel_data["sheets"]:
            all_text.append(f"\n{'=' * 80}")
            all_text.append(f"SHEET: {sheet['name']}")
            all_text.append(f"{'=' * 80}")
            all_text.append(f"Dimensions: {sheet['summary']['total_rows']} rows × {sheet['summary']['total_cols']} cols")
            all_text.append(f"Non-empty cells: {sheet['summary']['non_empty_cells']}")
            
            # Merged cells
            if "merged_cells" in sheet:
                all_text.append(f"\nMERGED CELLS: {len(sheet['merged_cells'])}")
                for merge in sheet["merged_cells"]:
                    all_text.append(f"  {merge['range']}")
            
            # Images
            if "images" in sheet:
                all_text.append(f"\nIMAGES: {len(sheet['images'])}")
                for img in sheet["images"]:
                    all_text.append(f"  Position: {img['anchor']}, Size: {img['width']}x{img['height']}")
            
            # Charts
            if "charts" in sheet:
                all_text.append(f"\nCHARTS: {len(sheet['charts'])}")
                for chart in sheet["charts"]:
                    all_text.append(f"  Type: {chart['type']}, Title: {chart.get('title', 'N/A')}")
            
            all_text.append(f"\nCELL DATA:")
            all_text.append("-" * 80)
            
            # Cell data
            for cell_ref, cell_data in sorted(sheet['data'].items()):
                value = cell_data[0].replace('\n', '\\n')
                data_type = cell_data[1]
                
                line = f"{cell_ref:6} = {value:50} [{data_type}]"
                
                # Add formatting info if present
                if len(cell_data) > 2:
                    fmt = cell_data[2]
                    fmt_parts = []
                    if "font" in fmt and fmt["font"].get("bold"):
                        fmt_parts.append("BOLD")
                    if "font" in fmt and fmt["font"].get("italic"):
                        fmt_parts.append("ITALIC")
                    if "fill" in fmt:
                        fmt_parts.append(f"BG:{fmt['fill']['fgColor'][:6]}")
                    
                    if fmt_parts:
                        line += f" <{', '.join(fmt_parts)}>"
                
                all_text.append(line)
            
            all_text.append("")
        
        return "\n".join(all_text)
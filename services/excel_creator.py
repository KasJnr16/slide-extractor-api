"""
Advanced Excel creation services - creates Excel files from extracted data
Supports: merged cells, images, charts, formatting, formulas, and more
"""
import io
import base64
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart,
    Reference, Series
)
from typing import Dict, Any, Optional
from datetime import datetime
import pandas as pd
from PIL import Image as PILImage


class ExcelCreator:
    """
    Creates Excel files from extracted data structure with full fidelity.
    Recreates: merged cells, images, charts, formatting, formulas.
    """
    
    def __init__(self):
        pass
    
    def _is_valid_color(self, color: str) -> bool:
        """
        Validate if a color string is in valid ARGB hex format.
        
        Args:
            color: Color string to validate
            
        Returns:
            True if valid, False otherwise
        """
        if not color or not isinstance(color, str):
            return False
        
        # Must be 8 hex digits (ARGB)
        if len(color) != 8:
            return False
        
        try:
            int(color, 16)
            return True
        except ValueError:
            return False
    
    def create_from_extracted_data(self, extracted_data: Dict[str, Any], 
                                   output_filename: Optional[str] = None) -> bytes:
        """
        Create Excel file from extracted data structure.
        
        Args:
            extracted_data: Dictionary from ExcelExtractor.extract_from_bytes()
            output_filename: Optional filename to save directly
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            # Create a new workbook
            workbook = openpyxl.Workbook()
            
            # Remove the default sheet
            workbook.remove(workbook.active)
            
            # Recreate each sheet
            for sheet_data in extracted_data["sheets"]:
                self._create_sheet(workbook, sheet_data)
            
            # Apply workbook metadata if present
            if "metadata" in extracted_data:
                self._apply_workbook_metadata(workbook, extracted_data["metadata"])
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            excel_bytes = output.getvalue()
            
            # Save to file if requested
            if output_filename:
                self.save_to_file(excel_bytes, output_filename)
            
            return excel_bytes
            
        except Exception as e:
            raise ValueError(f"Error creating Excel file: {str(e)}")
    
    def _create_sheet(self, workbook: openpyxl.Workbook, sheet_data: Dict[str, Any]):
        """
        Create a single sheet from extracted data with all features.
        
        Args:
            workbook: openpyxl workbook object
            sheet_data: Sheet data dictionary from extracted format
        """
        # Create sheet
        sheet = workbook.create_sheet(title=sheet_data["name"])
        
        # 1. Populate cell data with values and formatting
        if "data" in sheet_data and isinstance(sheet_data["data"], dict):
            for cell_ref, cell_data in sheet_data["data"].items():
                # Parse cell reference
                col_letter = ''.join([c for c in cell_ref if c.isalpha()])
                row_number = int(''.join([c for c in cell_ref if c.isdigit()]))
                col_index = column_index_from_string(col_letter)
                
                # Get cell
                cell = sheet.cell(row=row_number, column=col_index)
                
                # Extract data components
                value = cell_data[0]
                data_type = cell_data[1] if len(cell_data) > 1 else "string"
                formatting = cell_data[2] if len(cell_data) > 2 else None
                
                # Set cell value
                if str(value).strip():
                    converted_value = self._convert_value(str(value), data_type)
                    cell.value = converted_value
                
                # Apply formatting
                if formatting:
                    self._apply_cell_formatting(cell, formatting)
        
        # 2. Apply merged cells
        if "merged_cells" in sheet_data:
            for merge_info in sheet_data["merged_cells"]:
                try:
                    sheet.merge_cells(merge_info["range"])
                except Exception as e:
                    print(f"Warning: Could not merge cells {merge_info['range']}: {e}")
        
        # 3. Apply column widths
        if "column_widths" in sheet_data:
            for col_letter, width in sheet_data["column_widths"].items():
                try:
                    sheet.column_dimensions[col_letter].width = width
                except Exception as e:
                    print(f"Warning: Could not set column width for {col_letter}: {e}")
        
        # 4. Apply row heights
        if "row_heights" in sheet_data:
            for row_idx, height in sheet_data["row_heights"].items():
                try:
                    sheet.row_dimensions[row_idx].height = height
                except Exception as e:
                    print(f"Warning: Could not set row height for {row_idx}: {e}")
        
        # 5. Add images
        if "images" in sheet_data:
            for img_info in sheet_data["images"]:
                try:
                    self._add_image(sheet, img_info)
                except Exception as e:
                    print(f"Warning: Could not add image: {e}")
        
        # 6. Add charts
        if "charts" in sheet_data:
            for chart_info in sheet_data["charts"]:
                try:
                    self._add_chart(sheet, chart_info)
                except Exception as e:
                    print(f"Warning: Could not add chart: {e}")
        
        # 7. Apply conditional formatting
        if "conditional_formatting" in sheet_data:
            for cond_fmt in sheet_data["conditional_formatting"]:
                try:
                    self._add_conditional_formatting(sheet, cond_fmt)
                except Exception as e:
                    print(f"Warning: Could not add conditional formatting: {e}")
        
        # 8. Apply data validations
        if "data_validations" in sheet_data:
            for validation in sheet_data["data_validations"]:
                try:
                    self._add_data_validation(sheet, validation)
                except Exception as e:
                    print(f"Warning: Could not add data validation: {e}")
    
    def _convert_value(self, value: str, data_type: str) -> Any:
        """
        Convert string value back to appropriate data type.
        
        Args:
            value: String value from extracted data
            data_type: Original data type
            
        Returns:
            Converted value
        """
        if data_type == "number":
            try:
                # Try to convert to float first, then int if it's whole
                float_val = float(value)
                if float_val.is_integer():
                    return int(float_val)
                return float_val
            except ValueError:
                return value
        
        elif data_type == "date":
            try:
                # Try to parse date string
                return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    # Try alternative date formats
                    return pd.to_datetime(value).to_pydatetime()
                except:
                    return value
        
        elif data_type == "boolean":
            return value.lower() in ['true', '1', 'yes', 'on']
        
        elif data_type == "formula":
            # Return as formula (should start with =)
            if not value.startswith('='):
                return f"={value}"
            return value
        
        else:
            return value
    
    def _apply_cell_formatting(self, cell, formatting: Dict[str, Any]):
        """
        Apply formatting to a cell.
        
        Args:
            cell: openpyxl cell object
            formatting: Formatting dictionary from extracted data
        """
        try:
            # Apply font formatting
            if "font" in formatting:
                font_data = formatting["font"]
                
                # Validate color format (must be ARGB hex)
                color = font_data.get("color")
                if color and not self._is_valid_color(color):
                    color = None
                
                font = Font(
                    name=font_data.get("name", "Calibri"),
                    size=font_data.get("size", 11),
                    bold=font_data.get("bold", False),
                    italic=font_data.get("italic", False),
                    underline=font_data.get("underline", "none"),
                    color=color
                )
                cell.font = font
            
            # Apply fill/background color
            if "fill" in formatting:
                fill_data = formatting["fill"]
                fg_color = fill_data.get("fgColor")
                
                # Validate color format
                if fg_color and self._is_valid_color(fg_color):
                    fill = PatternFill(
                        start_color=fg_color,
                        end_color=fg_color,
                        patternType=fill_data.get("patternType", "solid")
                    )
                    cell.fill = fill
            
            # Apply alignment
            if "alignment" in formatting:
                align_data = formatting["alignment"]
                alignment = Alignment(
                    horizontal=align_data.get("horizontal"),
                    vertical=align_data.get("vertical"),
                    wrap_text=align_data.get("wrap_text", False),
                    text_rotation=align_data.get("text_rotation", 0)
                )
                cell.alignment = alignment
            
            # Apply border
            if "border" in formatting:
                border_data = formatting["border"]
                sides = {}
                
                for side_name in ['left', 'right', 'top', 'bottom']:
                    if side_name in border_data:
                        side_info = border_data[side_name]
                        # Validate border color
                        border_color = side_info.get("color")
                        if border_color and not self._is_valid_color(border_color):
                            border_color = None
                        
                        sides[side_name] = Side(
                            style=side_info.get("style", "thin"),
                            color=border_color
                        )
                
                if sides:
                    border = Border(**sides)
                    cell.border = border
            
            # Apply number format
            if "number_format" in formatting:
                cell.number_format = formatting["number_format"]
                
        except Exception as e:
            print(f"Warning: Could not apply formatting to cell: {e}")
    
    def _add_image(self, sheet, img_info: Dict[str, Any]):
        """
        Add an image to the sheet.
        
        Args:
            sheet: openpyxl sheet object
            img_info: Image information dictionary
        """
        if "data" not in img_info:
            return
        
        try:
            # Decode base64 image
            img_bytes = base64.b64decode(img_info["data"])
            
            # Create PIL Image
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            
            # Save to bytes
            img_output = io.BytesIO()
            pil_img.save(img_output, format=img_info.get("format", "PNG"))
            img_output.seek(0)
            
            # Create Excel image
            xl_img = XLImage(img_output)
            
            # Set dimensions if available
            if "width" in img_info:
                xl_img.width = img_info["width"]
            if "height" in img_info:
                xl_img.height = img_info["height"]
            
            # Add to sheet at anchor position
            anchor = img_info.get("anchor", "A1")
            sheet.add_image(xl_img, anchor)
            
        except Exception as e:
            raise ValueError(f"Error adding image: {e}")
    
    def _add_chart(self, sheet, chart_info: Dict[str, Any]):
        """
        Add a chart to the sheet.
        
        Args:
            sheet: openpyxl sheet object
            chart_info: Chart information dictionary
        """
        # Map chart type names to classes
        chart_classes = {
            "BarChart": BarChart,
            "LineChart": LineChart,
            "PieChart": PieChart,
            "AreaChart": AreaChart,
            "ScatterChart": ScatterChart
        }
        
        chart_type = chart_info.get("type", "BarChart")
        ChartClass = chart_classes.get(chart_type, BarChart)
        
        try:
            # Create chart
            chart = ChartClass()
            
            # Set title
            if "title" in chart_info:
                chart.title = chart_info["title"]
            
            # Set axis titles
            if "x_axis_title" in chart_info:
                chart.x_axis.title = chart_info["x_axis_title"]
            if "y_axis_title" in chart_info:
                chart.y_axis.title = chart_info["y_axis_title"]
            
            # Add series data
            if "series" in chart_info:
                for series_data in chart_info["series"]:
                    try:
                        # Parse reference strings back to Reference objects
                        if "values" in series_data and series_data["values"]:
                            # This is simplified - may need more robust parsing
                            values_ref = self._parse_reference(sheet, series_data["values"])
                            
                            if "categories" in series_data and series_data["categories"]:
                                cats_ref = self._parse_reference(sheet, series_data["categories"])
                                chart.add_data(values_ref, titles_from_data=True)
                                chart.set_categories(cats_ref)
                            else:
                                chart.add_data(values_ref, titles_from_data=True)
                    except Exception as e:
                        print(f"Warning: Could not add chart series: {e}")
            
            # Add chart to sheet
            anchor = chart_info.get("anchor", "E5")
            sheet.add_chart(chart, anchor)
            
        except Exception as e:
            raise ValueError(f"Error adding chart: {e}")
    
    def _parse_reference(self, sheet, ref_string: str):
        """
        Parse a reference string to a Reference object.
        
        Args:
            sheet: openpyxl sheet object
            ref_string: Reference string like "Sheet1!$A$1:$A$10"
            
        Returns:
            Reference object
        """
        try:
            # Simple parsing - extract range
            # Format: "Sheet1!$A$1:$A$10" or just "$A$1:$A$10"
            if "!" in ref_string:
                ref_string = ref_string.split("!")[1]
            
            # Remove $ signs
            ref_string = ref_string.replace("$", "")
            
            # Parse range
            if ":" in ref_string:
                start, end = ref_string.split(":")
                start_col = ''.join([c for c in start if c.isalpha()])
                start_row = int(''.join([c for c in start if c.isdigit()]))
                end_col = ''.join([c for c in end if c.isalpha()])
                end_row = int(''.join([c for c in end if c.isdigit()]))
                
                return Reference(
                    sheet,
                    min_col=column_index_from_string(start_col),
                    min_row=start_row,
                    max_col=column_index_from_string(end_col),
                    max_row=end_row
                )
            else:
                # Single cell reference
                col = ''.join([c for c in ref_string if c.isalpha()])
                row = int(''.join([c for c in ref_string if c.isdigit()]))
                
                return Reference(
                    sheet,
                    min_col=column_index_from_string(col),
                    min_row=row,
                    max_col=column_index_from_string(col),
                    max_row=row
                )
        except Exception as e:
            print(f"Warning: Could not parse reference {ref_string}: {e}")
            # Return a default reference
            return Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=1)
    
    def _add_conditional_formatting(self, sheet, cond_fmt: Dict[str, Any]):
        """
        Add conditional formatting to the sheet.
        
        Args:
            sheet: openpyxl sheet object
            cond_fmt: Conditional formatting dictionary
        """
        try:
            from openpyxl.formatting.rule import Rule
            
            # This is a simplified implementation
            # Full conditional formatting recreation is complex
            # and may require more specific handling per rule type
            
            if "range" in cond_fmt and "formula" in cond_fmt:
                rule = Rule(
                    type=cond_fmt.get("type", "expression"),
                    formula=[cond_fmt["formula"]],
                    priority=cond_fmt.get("priority", 1)
                )
                sheet.conditional_formatting.add(cond_fmt["range"], rule)
                
        except Exception as e:
            print(f"Warning: Could not add conditional formatting: {e}")
    
    def _add_data_validation(self, sheet, validation: Dict[str, Any]):
        """
        Add data validation to the sheet.
        
        Args:
            sheet: openpyxl sheet object
            validation: Data validation dictionary
        """
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            
            # Create data validation
            dv = DataValidation(
                type=validation.get("type", "list"),
                formula1=validation.get("formula1"),
                formula2=validation.get("formula2"),
                allow_blank=validation.get("allow_blank", True),
                showDropDown=validation.get("show_dropdown", True)
            )
            
            # Add ranges
            if "ranges" in validation:
                for range_str in validation["ranges"]:
                    dv.add(range_str)
            
            sheet.add_data_validation(dv)
            
        except Exception as e:
            print(f"Warning: Could not add data validation: {e}")
    
    def _apply_workbook_metadata(self, workbook, metadata: Dict[str, Any]):
        """
        Apply workbook-level metadata.
        
        Args:
            workbook: openpyxl workbook object
            metadata: Metadata dictionary
        """
        try:
            if "creator" in metadata:
                workbook.properties.creator = metadata["creator"]
            if "title" in metadata:
                workbook.properties.title = metadata["title"]
            if "subject" in metadata:
                workbook.properties.subject = metadata["subject"]
            if "created" in metadata:
                try:
                    workbook.properties.created = datetime.fromisoformat(metadata["created"])
                except:
                    pass
            if "modified" in metadata:
                try:
                    workbook.properties.modified = datetime.fromisoformat(metadata["modified"])
                except:
                    pass
        except Exception as e:
            print(f"Warning: Could not apply workbook metadata: {e}")
    
    def save_to_file(self, excel_bytes: bytes, filename: str):
        """
        Save Excel bytes to file.
        
        Args:
            excel_bytes: Excel file content as bytes
            filename: Output filename
        """
        try:
            with open(filename, 'wb') as f:
                f.write(excel_bytes)
            print(f"✅ Excel file saved as: {filename}")
        except Exception as e:
            raise ValueError(f"Error saving file: {str(e)}")
    
    def create_excel_from_json(self, json_data: str, output_filename: Optional[str] = None) -> bytes:
        """
        Create Excel file from JSON string.
        
        Args:
            json_data: JSON string containing extracted data
            output_filename: Optional filename to save directly
            
        Returns:
            bytes: Excel file content as bytes
        """
        import json
        
        try:
            data = json.loads(json_data)
            return self.create_from_extracted_data(data, output_filename)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON data: {str(e)}")


# Utility functions for quick Excel creation
def create_excel_from_extracted_data(extracted_data: Dict[str, Any], 
                                     output_filename: Optional[str] = None) -> bytes:
    """
    Quick utility function to create Excel from extracted data.
    
    Args:
        extracted_data: Data from ExcelExtractor
        output_filename: Optional filename to save directly
        
    Returns:
        bytes: Excel file content
    """
    creator = ExcelCreator()
    return creator.create_from_extracted_data(extracted_data, output_filename)


def create_excel_from_json(json_data: str, output_filename: Optional[str] = None) -> bytes:
    """
    Quick utility function to create Excel from JSON.
    
    Args:
        json_data: JSON string containing extracted data
        output_filename: Optional filename to save directly
        
    Returns:
        bytes: Excel file content
    """
    creator = ExcelCreator()
    return creator.create_excel_from_json(json_data, output_filename)
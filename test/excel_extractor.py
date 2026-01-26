import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import io
from typing import List, Dict, Any, Optional
from datetime import datetime

class ExcelTableExtractor:
    """
    Extracts text from Excel files with row and column references for easy matching.
    Supports both .xlsx and .xls formats.
    """
    
    def __init__(self):
        pass
    
    def extract_from_bytes(self, file_bytes: bytes, filename: str) -> Dict[str, Any]:
        """
        Extract Excel content with row/column references from bytes.
        
        Args:
            file_bytes: Excel file content as bytes
            filename: Name of the file (for format detection)
            
        Returns:
            Dictionary containing extracted data with structure:
            {
                "filename": str,
                "sheets": [
                    {
                        "name": str,
                        "data": [
                            {
                                "cell": str,  # e.g., "A1", "B2"
                                "row": int,
                                "col": int,
                                "col_letter": str,
                                "value": str,
                                "data_type": str  # "string", "number", "date", "formula", "empty"
                            }
                        ],
                        "summary": {
                            "total_rows": int,
                            "total_cols": int,
                            "non_empty_cells": int
                        }
                    }
                ]
            }
        """
        try:
            if filename.lower().endswith(('.xlsx', '.xlsm')):
                return self._extract_xlsx(file_bytes)
            elif filename.lower().endswith('.xls'):
                return self._extract_xls(file_bytes)
            else:
                raise ValueError(f"Unsupported Excel format: {filename}")
        except Exception as e:
            raise ValueError(f"Error extracting Excel data: {str(e)}")
    
    def _extract_xlsx(self, file_bytes: bytes) -> Dict[str, Any]:
        """Extract from modern Excel format (.xlsx) using openpyxl."""
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        sheets_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            non_empty_count = 0
            
            # Get actual used range
            max_row = sheet.max_row or 1
            max_col = sheet.max_column or 1
            
            # Extract all cells in the used range
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = self._get_cell_value(cell)
                    data_type = self._get_cell_type(cell)
                    
                    # Convert value to string properly
                    if value is None:
                        value_str = ""
                    elif isinstance(value, datetime):
                        value_str = value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        value_str = str(value)
                    
                    if value_str.strip():
                        non_empty_count += 1
                    
                    cell_info = {
                        "cell": f"{get_column_letter(col_idx)}{row_idx}",
                        "row": row_idx,
                        "col": col_idx,
                        "col_letter": get_column_letter(col_idx),
                        "value": value_str,
                        "data_type": data_type
                    }
                    sheet_data.append(cell_info)
            
            sheets_data.append({
                "name": sheet_name,
                "data": sheet_data,
                "summary": {
                    "total_rows": max_row,
                    "total_cols": max_col,
                    "non_empty_cells": non_empty_count
                }
            })
        
        return {
            "filename": "excel_file",
            "sheets": sheets_data
        }
    
    def _extract_xls(self, file_bytes: bytes) -> Dict[str, Any]:
        """Extract from legacy Excel format (.xls) using pandas."""
        excel_file = io.BytesIO(file_bytes)
        
        try:
            # Read all sheets
            xls = pd.ExcelFile(excel_file, engine='xlrd')
            sheets_data = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
                sheet_data = []
                non_empty_count = 0
                
                for row_idx in range(len(df)):
                    for col_idx in range(len(df.columns)):
                        value = df.iloc[row_idx, col_idx]
                        
                        # Convert value to string properly
                        if pd.isna(value):
                            value_str = ""
                        elif isinstance(value, datetime):
                            value_str = value.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            value_str = str(value)
                        
                        if value_str.strip():
                            non_empty_count += 1
                        
                        cell_info = {
                            "cell": f"{get_column_letter(col_idx + 1)}{row_idx + 1}",
                            "row": row_idx + 1,
                            "col": col_idx + 1,
                            "col_letter": get_column_letter(col_idx + 1),
                            "value": value_str,
                            "data_type": self._infer_data_type(value)
                        }
                        sheet_data.append(cell_info)
                
                sheets_data.append({
                    "name": sheet_name,
                    "data": sheet_data,
                    "summary": {
                        "total_rows": len(df),
                        "total_cols": len(df.columns),
                        "non_empty_cells": non_empty_count
                    }
                })
            
            return {
                "filename": "excel_file",
                "sheets": sheets_data
            }
            
        except Exception as e:
            raise ValueError(f"Error reading .xls file: {str(e)}")
    
    def _get_cell_value(self, cell) -> Any:
        """Get the actual value from an openpyxl cell."""
        try:
            # Handle merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                return ""
            
            return cell.value
        except Exception:
            return ""
    
    def _get_cell_type(self, cell) -> str:
        """Determine the data type of a cell."""
        try:
            # Handle merged cells
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                return "merged"
            
            if cell.value is None:
                return "empty"
            
            if cell.data_type == 'f':
                return "formula"
            elif cell.is_date:
                return "date"
            elif cell.data_type == 'n':
                return "number"
            elif cell.data_type == 's':
                return "string"
            elif cell.data_type == 'b':
                return "boolean"
            else:
                return "empty"
        except Exception:
            return "unknown"
    
    def _infer_data_type(self, value) -> str:
        """Infer data type for pandas-based extraction."""
        if pd.isna(value) or value == "":
            return "empty"
        elif isinstance(value, bool):
            return "boolean"
        elif isinstance(value, (int, float)):
            return "number"
        elif isinstance(value, datetime):
            return "date"
        elif isinstance(value, str):
            # Try to detect if it's a date
            try:
                pd.to_datetime(value)
                return "date"
            except:
                return "string"
        else:
            return "unknown"
    
    def get_table_structure(self, extracted_data: Dict[str, Any], sheet_name: str = None) -> Dict[str, Any]:
        """
        Analyze the structure of extracted data to identify tables and patterns.
        
        Args:
            extracted_data: Data from extract_from_bytes
            sheet_name: Specific sheet to analyze (optional)
            
        Returns:
            Dictionary with table structure analysis
        """
        if sheet_name:
            sheets_to_analyze = [s for s in extracted_data["sheets"] if s["name"] == sheet_name]
        else:
            sheets_to_analyze = extracted_data["sheets"]
        
        analysis = {}
        
        for sheet in sheets_to_analyze:
            sheet_name = sheet["name"]
            data = sheet["data"]
            
            # Create a grid representation
            max_row = max((item["row"] for item in data), default=0)
            max_col = max((item["col"] for item in data), default=0)
            
            if max_row == 0 or max_col == 0:
                analysis[sheet_name] = {
                    "tables": [],
                    "total_rows": 0,
                    "total_cols": 0,
                    "non_empty_cells": 0
                }
                continue
            
            grid = [[None for _ in range(max_col)] for _ in range(max_row)]
            
            for item in data:
                if item["value"].strip():
                    grid[item["row"] - 1][item["col"] - 1] = item
            
            # Identify potential tables (contiguous blocks of non-empty cells)
            tables = self._identify_tables(grid)
            
            analysis[sheet_name] = {
                "tables": tables,
                "total_rows": max_row,
                "total_cols": max_col,
                "non_empty_cells": sheet["summary"]["non_empty_cells"]
            }
        
        return analysis
    
    def _identify_tables(self, grid: List[List[Optional[Dict]]]) -> List[Dict[str, Any]]:
        """Identify contiguous blocks of data as tables."""
        if not grid or not grid[0]:
            return []
        
        rows = len(grid)
        cols = len(grid[0])
        visited = [[False for _ in range(cols)] for _ in range(rows)]
        tables = []
        
        for i in range(rows):
            for j in range(cols):
                if grid[i][j] is not None and not visited[i][j]:
                    # Found start of a new table
                    table_bounds = self._find_table_bounds(grid, visited, i, j)
                    if table_bounds:
                        tables.append(table_bounds)
        
        return tables
    
    def _find_table_bounds(self, grid: List[List[Optional[Dict]]], visited: List[List[bool]], 
                          start_row: int, start_col: int) -> Optional[Dict[str, Any]]:
        """Find the bounds of a contiguous table starting from a cell."""
        rows = len(grid)
        cols = len(grid[0])
        
        # Use iterative flood fill to avoid recursion depth issues
        stack = [(start_row, start_col)]
        min_row, max_row = start_row, start_row
        min_col, max_col = start_col, start_col
        cells = []
        
        while stack:
            row, col = stack.pop()
            
            # Check bounds and if already visited
            if (row < 0 or row >= rows or col < 0 or col >= cols or 
                visited[row][col]):
                continue
            
            # Check if cell has data
            if grid[row][col] is None:
                continue
            
            # Mark as visited and add to cells
            visited[row][col] = True
            cells.append(grid[row][col])
            
            # Update bounds
            min_row = min(min_row, row)
            max_row = max(max_row, row)
            min_col = min(min_col, col)
            max_col = max(max_col, col)
            
            # Check adjacent cells (4-directional connectivity for tables)
            # Also check diagonals for more flexible detection
            directions = [
                (-1, 0),   # up
                (1, 0),    # down
                (0, -1),   # left
                (0, 1),    # right
                (-1, -1),  # diagonal up-left
                (-1, 1),   # diagonal up-right
                (1, -1),   # diagonal down-left
                (1, 1)     # diagonal down-right
            ]
            
            for dr, dc in directions:
                new_row, new_col = row + dr, col + dc
                if (0 <= new_row < rows and 0 <= new_col < cols and 
                    not visited[new_row][new_col] and 
                    grid[new_row][new_col] is not None):
                    stack.append((new_row, new_col))
        
        # Only consider it a table if it has multiple cells
        if len(cells) < 2:
            return None
        
        return {
            "start_cell": f"{get_column_letter(min_col + 1)}{min_row + 1}",
            "end_cell": f"{get_column_letter(max_col + 1)}{max_row + 1}",
            "start_row": min_row + 1,
            "end_row": max_row + 1,
            "start_col": min_col + 1,
            "end_col": max_col + 1,
            "cell_count": len(cells),
            "cells": cells
        }
    
    def display_as_table(self, extracted_data: Dict[str, Any], sheet_name: str = None):
        """
        Display extracted data in a table format (Excel-like view).
        
        Args:
            extracted_data: Data from extract_from_bytes
            sheet_name: Specific sheet to display (optional, displays all if None)
        """
        if sheet_name:
            sheets_to_display = [s for s in extracted_data["sheets"] if s["name"] == sheet_name]
        else:
            sheets_to_display = extracted_data["sheets"]
        
        for sheet in sheets_to_display:
            print(f"\n{'='*80}")
            print(f"📊 SHEET: {sheet['name']}")
            print(f"{'='*80}")
            print(f"Dimensions: {sheet['summary']['total_rows']} rows × {sheet['summary']['total_cols']} cols")
            print(f"Non-empty cells: {sheet['summary']['non_empty_cells']}")
            print(f"{'-'*80}\n")
            
            # Build a grid for display
            max_row = sheet['summary']['total_rows']
            max_col = sheet['summary']['total_cols']
            
            grid = [["" for _ in range(max_col)] for _ in range(max_row)]
            
            for cell in sheet['data']:
                if cell['value'].strip():
                    grid[cell['row'] - 1][cell['col'] - 1] = cell['value']
            
            # Create pandas DataFrame for nice table display
            df = pd.DataFrame(grid)
            df.columns = [get_column_letter(i+1) for i in range(max_col)]
            df.index = [i+1 for i in range(max_row)]
            
            print(df.to_string())
            print()


# Test function
def test_excel_extractor():
    """Test the Excel extractor with a user-selected Excel file."""
    import os
    import tkinter as tk
    from tkinter import filedialog
    
    # Hide the main tkinter window
    root = tk.Tk()
    root.withdraw()
    
    # Ask user to select an Excel file
    print("🔍 Please select an Excel file to test...")
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xls"),
            ("All files", "*.*")
        ]
    )
    
    if not file_path:
        print("❌ No file selected. Exiting.")
        return False
    
    print(f"📁 Selected file: {file_path}")
    
    # Read the selected file
    try:
        with open(file_path, 'rb') as f:
            file_bytes = f.read()
        filename = os.path.basename(file_path)
    except Exception as e:
        print(f"❌ Error reading file: {str(e)}")
        return False
    
    # Test extraction
    extractor = ExcelTableExtractor()
    
    try:
        result = extractor.extract_from_bytes(file_bytes, filename)
        
        print("✅ Excel extraction successful!")
        print(f"Found {len(result['sheets'])} sheets\n")
        
        # Ask user which format they want
        print("Choose display format:")
        print("1. Cell-by-cell (Key:Value pairs) - Best for AI processing")
        print("2. Table format (Excel-like view) - Best for human reading")
        print("3. Both formats")
        
        choice = input("Enter choice (1/2/3): ").strip()
        
        if choice in ['1', '3']:
            print("\n" + "="*80)
            print("📋 CELL-BY-CELL FORMAT (All Data)")
            print("="*80)
            
            for sheet in result['sheets']:
                print(f"\n📊 Sheet: {sheet['name']}")
                print(f"   Dimensions: {sheet['summary']['total_rows']} rows × {sheet['summary']['total_cols']} cols")
                print(f"   Non-empty cells: {sheet['summary']['non_empty_cells']}\n")
                
                # Show ALL non-empty cells
                non_empty_cells = [c for c in sheet['data'] if c['value'].strip()]
                
                for cell in non_empty_cells:
                    value_display = cell['value'].replace('\n', '\\n')  # Show newlines
                    print(f"   {cell['cell']:6} = {value_display} [{cell['data_type']}]")
                
                print()
        
        if choice in ['2', '3']:
            print("\n" + "="*80)
            print("📊 TABLE FORMAT")
            print("="*80)
            extractor.display_as_table(result)
        
        # Test table structure analysis
        print("\n" + "="*80)
        print("🔍 TABLE STRUCTURE ANALYSIS")
        print("="*80)
        structure = extractor.get_table_structure(result)
        
        for sheet_name, analysis in structure.items():
            print(f"\n📋 {sheet_name}:")
            print(f"   Tables found: {len(analysis['tables'])}")
            
            for i, table in enumerate(analysis['tables'], 1):
                print(f"   Table {i}: {table['start_cell']} to {table['end_cell']} ({table['cell_count']} cells)")
        
        return True
        
    except Exception as e:
        import traceback
        print(f"❌ Test failed: {str(e)}")
        print(traceback.format_exc())
        return False


def test_with_sample_data():
    """Test the Excel extractor with sample data (original test)."""
    import os
    
    # Create a sample Excel file in memory for testing
    sample_data = {
        "Sheet1": [
            ["Name", "Age", "City"],
            ["John Doe", 30, "New York"],
            ["Jane Smith", 25, "Los Angeles"],
            ["Bob Johnson", 35, "Chicago"]
        ],
        "Sheet2": [
            ["Product", "Price", "Stock"],
            ["Laptop", 999.99, 50],
            ["Mouse", 29.99, 200],
            ["Keyboard", 79.99, 100]
        ]
    }
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, data in sample_data.items():
            df = pd.DataFrame(data[1:], columns=data[0])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output.seek(0)
    
    # Test extraction
    extractor = ExcelTableExtractor()
    
    try:
        result = extractor.extract_from_bytes(output.getvalue(), "test.xlsx")
        
        print("✅ Excel extraction successful!")
        print(f"Found {len(result['sheets'])} sheets\n")
        
        # Show both formats
        print("="*80)
        print("📋 CELL-BY-CELL FORMAT")
        print("="*80)
        
        for sheet in result['sheets']:
            print(f"\n📊 Sheet: {sheet['name']}")
            print(f"   Dimensions: {sheet['summary']['total_rows']} rows × {sheet['summary']['total_cols']} cols")
            print(f"   Non-empty cells: {sheet['summary']['non_empty_cells']}\n")
            
            # Show ALL non-empty cells
            non_empty_cells = [c for c in sheet['data'] if c['value'].strip()]
            
            for cell in non_empty_cells:
                print(f"   {cell['cell']:6} = {cell['value']} [{cell['data_type']}]")
        
        print("\n" + "="*80)
        print("📊 TABLE FORMAT")
        print("="*80)
        extractor.display_as_table(result)
        
        # Test table structure analysis
        print("\n" + "="*80)
        print("🔍 TABLE STRUCTURE ANALYSIS")
        print("="*80)
        structure = extractor.get_table_structure(result)
        
        for sheet_name, analysis in structure.items():
            print(f"\n📋 {sheet_name}:")
            print(f"   Tables found: {len(analysis['tables'])}")
            
            for i, table in enumerate(analysis['tables'], 1):
                print(f"   Table {i}: {table['start_cell']} to {table['end_cell']} ({table['cell_count']} cells)")
        
        return True
        
    except Exception as e:
        import traceback
        print(f"❌ Test failed: {str(e)}")
        print(traceback.format_exc())
        return False


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import messagebox
    
    # Hide the main tkinter window
    root = tk.Tk()
    root.withdraw()
    
    # Ask user to choose test mode
    choice = messagebox.askyesno(
        "Test Mode Selection",
        "Do you want to select an Excel file from your computer?\n\n"
        "Yes = Select file\n"
        "No = Use sample data"
    )
    
    if choice:
        # User wants to select a file
        test_excel_extractor()
    else:
        # User wants to use sample data
        test_with_sample_data()